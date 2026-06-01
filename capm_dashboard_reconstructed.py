"""
CAPM Mutual Fund Research Workbench

Install:
  pip install streamlit pandas numpy yfinance mftool plotly openpyxl

Run:
  streamlit run capm_dashboard_reconstructed.py
"""

import html
import io
import os
import re
import tempfile
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import yfinance as yf
from mftool import Mftool
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
os.environ.setdefault("MPLCONFIGDIR", tempfile.gettempdir())


APP_TITLE = "CAPM Mutual Fund Research Workbench"
PAGES = ["Analyze Fund", "Compare Funds", "Forecast Planner", "Report Builder", "Formula Audit"]

PLOT_BG = "#FFFFFF"
PAGE_BG = "#F6F8FB"
INK = "#182230"
MUTED = "#667085"
BORDER = "#D9E2EC"
SOFT = "#EEF4F7"
NAVY = "#173B57"
TEAL = "#0F766E"
ORANGE = "#D97706"
BLUE = "#2563EB"
GREEN = "#15803D"
RED = "#B42318"
GOLD = "#A16207"
GRID = "#E8EEF3"


BENCHMARKS = {
    "Auto by category": {"ticker": "auto", "label": "Auto by category", "fallback": "^NSEI"},
    "Nifty 50": {"ticker": "^NSEI", "label": "Nifty 50", "fallback": "^BSESN"},
    "Nifty 500": {"ticker": "NSE500.NS", "label": "Nifty 500", "fallback": "^NSEI"},
    "Nifty Midcap 50": {"ticker": "^NSEMDCP50", "label": "Nifty Midcap 50", "fallback": "^NSEI"},
    "Nifty Midcap 150": {"ticker": "^NSMIDCP", "label": "Nifty Midcap 150", "fallback": "^NSEMDCP50"},
    "Nifty Smallcap 250": {"ticker": "NIFTYSMLCAP250.NS", "label": "Nifty Smallcap 250", "fallback": "^NSEI"},
    "Sensex": {"ticker": "^BSESN", "label": "Sensex", "fallback": "^NSEI"},
}

CATEGORY_BENCHMARK = {
    "LargeCap": "Nifty 50",
    "FlexiCap": "Nifty 500",
    "MidCap": "Nifty Midcap 150",
    "SmallCap": "Nifty Smallcap 250",
    "ELSS": "Nifty 50",
    "Debt": "Nifty 50",
    "Hybrid": "Nifty 50",
    "Index": "Nifty 50",
}

CATEGORY_KEYWORDS = {
    "LargeCap": ["large cap", "largecap", "bluechip", "top 100", "focused equity"],
    "MidCap": ["mid cap", "midcap"],
    "SmallCap": ["small cap", "smallcap"],
    "FlexiCap": ["flexi cap", "flexicap", "multi cap", "multicap"],
    "ELSS": ["elss", "tax saver", "tax plan"],
    "Debt": ["liquid", "overnight", "short duration", "corporate bond", "gilt"],
    "Hybrid": ["hybrid", "balanced advantage", "aggressive hybrid"],
    "Index": ["index", "nifty 50", "sensex"],
}


st.set_page_config(page_title="CAPM Research", page_icon=":chart_with_upwards_trend:", layout="wide")


st.markdown(
    f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=IBM+Plex+Mono:wght@400;500;600&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}
.stApp {{ background: {PAGE_BG}; color: {INK}; }}
[data-testid="stSidebar"], [data-testid="collapsedControl"] {{ display:none !important; }}
.block-container {{ padding-top: 1.4rem; max-width: 1240px; }}
h1, h2, h3 {{ letter-spacing: 0; color: {INK}; }}
a {{ color: {TEAL}; }}
.topbar {{
  display:flex; align-items:flex-end; justify-content:space-between; gap:24px;
  border-bottom:1px solid {BORDER}; padding-bottom:18px; margin-bottom:18px;
}}
.brand-title {{ font-size:26px; font-weight:800; line-height:1.12; color:{INK}; margin:0; }}
.brand-sub {{ font-size:13px; color:{MUTED}; margin-top:7px; line-height:1.5; }}
.status-pill {{
  display:inline-flex; align-items:center; gap:8px; padding:8px 12px; border-radius:999px;
  background:#FFFFFF; border:1px solid {BORDER}; color:{NAVY}; font-size:12px; font-weight:700;
}}
.page-note {{
  background:#FFFFFF; border:1px solid {BORDER}; border-left:4px solid {TEAL};
  border-radius:8px; padding:13px 15px; color:{MUTED}; font-size:13px; line-height:1.6;
}}
.section-title {{
  font-size:15px; font-weight:800; color:{INK}; margin:24px 0 12px 0; padding-bottom:9px;
  border-bottom:1px solid {BORDER};
}}
.panel {{
  background:#FFFFFF; border:1px solid {BORDER}; border-radius:8px;
  padding:16px 18px; box-shadow:0 10px 26px rgba(16,24,40,0.045);
}}
.panel-tight {{ padding:12px 14px; }}
.metric-card {{
  background:#FFFFFF; border:1px solid {BORDER}; border-radius:8px; padding:15px 16px;
  box-shadow:0 10px 26px rgba(16,24,40,0.045); min-height:92px;
}}
.metric-label {{ color:{MUTED}; font-size:11px; font-weight:800; text-transform:uppercase; letter-spacing:.06em; }}
.metric-value {{ color:{INK}; font-family:'IBM Plex Mono', monospace; font-size:22px; font-weight:700; margin-top:8px; }}
.metric-help {{ color:{MUTED}; font-size:12px; margin-top:5px; }}
.card-title {{
  color:{NAVY}; font-size:11px; font-weight:800; text-transform:uppercase; letter-spacing:.08em;
  margin-bottom:12px;
}}
.row {{
  display:flex; justify-content:space-between; gap:18px; align-items:flex-start;
  border-bottom:1px solid #EDF2F7; padding:8px 0; font-size:13px;
}}
.row:last-child {{ border-bottom:none; }}
.row span:first-child {{ color:{MUTED}; }}
.row span:last-child {{ color:{INK}; font-family:'IBM Plex Mono', monospace; font-weight:600; text-align:right; }}
.good {{ color:{GREEN} !important; }}
.bad {{ color:{RED} !important; }}
.warn {{ color:{GOLD} !important; }}
.tag {{
  display:inline-flex; align-items:center; padding:4px 8px; border-radius:999px;
  background:{SOFT}; color:{NAVY}; font-size:11px; font-weight:800;
}}
.stButton > button, .stDownloadButton > button {{
  border-radius:8px !important; border:1px solid {NAVY} !important; background:{NAVY} !important;
  color:white !important; font-weight:800 !important; padding:11px 16px !important;
}}
.stButton > button:hover, .stDownloadButton > button:hover {{ background:#0F2C42 !important; border-color:#0F2C42 !important; }}
[data-testid="stMetric"] {{
  background:#FFFFFF; border:1px solid {BORDER}; border-radius:8px; padding:14px 16px;
  box-shadow:0 10px 26px rgba(16,24,40,0.045);
}}
[data-testid="stMetricLabel"] {{ color:{MUTED}; font-size:11px; font-weight:800; text-transform:uppercase; }}
[data-testid="stMetricValue"] {{ color:{INK}; font-family:'IBM Plex Mono', monospace; font-size:22px; }}
.stTabs [data-baseweb="tab-list"] {{ gap:8px; }}
.stTabs [data-baseweb="tab"] {{
  background:#FFFFFF; border:1px solid {BORDER}; border-radius:8px; padding:10px 14px;
}}
.stRadio [role="radiogroup"] {{
  background:#FFFFFF; border:1px solid {BORDER}; border-radius:10px; padding:6px;
  box-shadow:0 10px 26px rgba(16,24,40,0.035);
}}
.stRadio [role="radio"] {{ padding:6px 10px; }}
.stDataFrame {{ border:1px solid {BORDER}; border-radius:8px; overflow:hidden; }}
@media (max-width: 760px) {{
  .topbar {{ display:block; }}
  .brand-title {{ font-size:21px; }}
  .block-container {{ padding-left:1rem; padding-right:1rem; }}
  .metric-value {{ font-size:18px; }}
}}
</style>
""",
    unsafe_allow_html=True,
)


def esc(value):
    return html.escape("" if value is None else str(value))


def pct(value, digits=2):
    try:
        value = float(value)
        if np.isnan(value) or np.isinf(value):
            return "N/A"
        return f"{value:.{digits}%}"
    except Exception:
        return "N/A"


def num(value, digits=3):
    try:
        value = float(value)
        if np.isnan(value) or np.isinf(value):
            return "N/A"
        return f"{value:.{digits}f}"
    except Exception:
        return "N/A"


def money(value):
    try:
        value = float(value)
        if np.isnan(value) or np.isinf(value):
            return "N/A"
        return "INR " + format(round(value), ",")
    except Exception:
        return "N/A"


def signed_class(value):
    try:
        value = float(value)
        return "good" if value > 0 else "bad" if value < 0 else "warn"
    except Exception:
        return "warn"


def row_html(label, value, cls=""):
    return f"<div class='row'><span>{esc(label)}</span><span class='{cls}'>{esc(value)}</span></div>"


def metric_card(label, value, help_text="", cls=""):
    return (
        "<div class='metric-card'>"
        f"<div class='metric-label'>{esc(label)}</div>"
        f"<div class='metric-value {cls}'>{esc(value)}</div>"
        f"<div class='metric-help'>{esc(help_text)}</div>"
        "</div>"
    )


def section(title):
    st.markdown(f"<div class='section-title'>{esc(title)}</div>", unsafe_allow_html=True)


def category_from_name(name):
    text = name.lower()
    if "small" in text:
        return "SmallCap"
    if "mid" in text and "small" not in text:
        return "MidCap"
    if "flexi" in text or "multi cap" in text or "multicap" in text:
        return "FlexiCap"
    if "elss" in text or "tax saver" in text:
        return "ELSS"
    if any(word in text for word in ["liquid", "overnight", "debt", "gilt", "bond"]):
        return "Debt"
    if any(word in text for word in ["hybrid", "balanced", "aggressive"]):
        return "Hybrid"
    if "index" in text or "nifty 50" in text or "sensex" in text:
        return "Index"
    return "LargeCap"


def category_label(category):
    return {
        "LargeCap": "Large Cap",
        "FlexiCap": "Flexi Cap",
        "MidCap": "Mid Cap",
        "SmallCap": "Small Cap",
        "ELSS": "ELSS",
        "Debt": "Debt",
        "Hybrid": "Hybrid",
        "Index": "Index",
    }.get(category, category)


def periods_per_year(frequency):
    return (252, "B") if frequency == "Daily" else (12, "M")


@st.cache_data(show_spinner=False, ttl=24 * 3600)
def load_schemes():
    mf = Mftool()
    data = mf.get_scheme_codes()
    df = pd.DataFrame(list(data.items()), columns=["code", "name"])
    df["code"] = df["code"].astype(str)
    df = df[df["code"].str.fullmatch(r"\d+")]
    df["name"] = df["name"].astype(str).str.strip()
    df["name_lower"] = df["name"].str.lower()
    return df.sort_values("name")


@st.cache_data(show_spinner=False, ttl=6 * 3600)
def fetch_fund_nav(scheme_code):
    mf = Mftool()
    raw = mf.get_scheme_historical_nav(str(scheme_code), as_Dataframe=True)
    if raw is None or raw.empty:
        return pd.Series(dtype=float)
    raw.index = pd.to_datetime(raw.index, format="%d-%m-%Y", errors="coerce")
    raw = raw.dropna().sort_index()
    nav = pd.to_numeric(raw["nav"], errors="coerce").dropna()
    nav.name = "fund"
    return nav


@st.cache_data(show_spinner=False, ttl=6 * 3600)
def fetch_live_nav(scheme_code):
    nav_series = fetch_fund_nav(str(scheme_code))
    if nav_series.empty:
        return None, "", 0.0, 0.0
    latest = float(nav_series.iloc[-1])
    latest_date = nav_series.index[-1].strftime("%d-%m-%Y")
    if len(nav_series) < 2:
        return latest, latest_date, 0.0, 0.0
    prev = float(nav_series.iloc[-2])
    change = latest - prev
    change_pct = change / prev if prev else 0.0
    return latest, latest_date, change, change_pct


@st.cache_data(show_spinner=False, ttl=6 * 3600)
def fetch_yahoo_series(ticker, start_date, end_date):
    data = yf.download(
        ticker,
        start=start_date,
        end=end_date,
        auto_adjust=True,
        progress=False,
        threads=False,
    )
    if data is None or data.empty:
        return pd.Series(dtype=float)
    close = data["Close"].squeeze()
    close = pd.to_numeric(close, errors="coerce").dropna()
    close.index = pd.to_datetime(close.index).tz_localize(None)
    close.name = "benchmark"
    return close.sort_index()


def read_uploaded_benchmark(uploaded_file):
    if uploaded_file is None:
        return pd.Series(dtype=float)
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
        date_col = next((c for c in df.columns if str(c).lower() in ["date", "nav date", "timestamp"]), df.columns[0])
        price_col = next(
            (c for c in df.columns if str(c).lower() in ["close", "price", "nav", "index", "value"]),
            df.columns[-1],
        )
        series = pd.Series(pd.to_numeric(df[price_col], errors="coerce").values, index=pd.to_datetime(df[date_col], errors="coerce"))
        series = series.dropna().sort_index()
        series.name = "benchmark"
        return series
    except Exception:
        return pd.Series(dtype=float)


def resolve_benchmark(category, benchmark_choice):
    if benchmark_choice == "Auto by category":
        benchmark_choice = CATEGORY_BENCHMARK.get(category, "Nifty 50")
    details = BENCHMARKS.get(benchmark_choice, BENCHMARKS["Nifty 50"])
    return details["label"], details["ticker"], details.get("fallback", "^NSEI")


def fetch_benchmark_prices(category, benchmark_choice, start_date, end_date, uploaded_series=None):
    if uploaded_series is not None and not uploaded_series.empty:
        filtered = uploaded_series[(uploaded_series.index >= start_date) & (uploaded_series.index <= end_date)]
        return filtered.rename("benchmark"), "Uploaded benchmark", "custom"

    label, ticker, fallback = resolve_benchmark(category, benchmark_choice)
    market = fetch_yahoo_series(ticker, start_date.strftime("%Y-%m-%d"), (end_date + timedelta(days=1)).strftime("%Y-%m-%d"))
    source = ticker
    if market.empty and fallback:
        market = fetch_yahoo_series(fallback, start_date.strftime("%Y-%m-%d"), (end_date + timedelta(days=1)).strftime("%Y-%m-%d"))
        source = f"{fallback} fallback"
    return market.rename("benchmark"), label, source


def return_frame_from_prices(fund_prices, benchmark_prices, frequency):
    prices = pd.concat([fund_prices.rename("fund"), benchmark_prices.rename("benchmark")], axis=1).dropna()
    prices = prices.sort_index()
    if frequency == "Monthly":
        prices = prices.resample("M").last().dropna()
    returns = prices.pct_change().dropna()
    returns.columns = ["fund_return", "benchmark_return"]
    return prices, returns


def calculate_drawdown(series):
    if series.empty:
        return pd.Series(dtype=float)
    indexed = series / series.iloc[0] * 10000
    return (indexed - indexed.cummax()) / indexed.cummax()


def cagr(series):
    if len(series) < 2:
        return np.nan
    years = max((series.index[-1] - series.index[0]).days / 365.25, 1 / 365.25)
    return (float(series.iloc[-1]) / float(series.iloc[0])) ** (1 / years) - 1


def run_capm_analysis(scheme_code, fund_name, category, benchmark_choice, years, frequency, rf_annual, uploaded_benchmark=None):
    as_of = pd.Timestamp.today().normalize()
    start_date = as_of - timedelta(days=int(years * 365.25))
    periods, _ = periods_per_year(frequency)

    fund_all = fetch_fund_nav(str(scheme_code))
    if fund_all.empty:
        return None, "No AMFI NAV history was found for this fund."

    fund_prices = fund_all[(fund_all.index >= start_date) & (fund_all.index <= as_of)]
    if len(fund_prices) < (45 if frequency == "Daily" else 12):
        return None, "Not enough NAV observations for the selected period."

    uploaded_series = read_uploaded_benchmark(uploaded_benchmark)
    benchmark_prices, benchmark_label, benchmark_source = fetch_benchmark_prices(category, benchmark_choice, start_date, as_of, uploaded_series)
    if benchmark_prices.empty:
        return None, "Benchmark data is unavailable. Try Nifty 50, a shorter period, or upload a benchmark CSV/XLSX."

    prices, returns = return_frame_from_prices(fund_prices, benchmark_prices, frequency)
    min_obs = 45 if frequency == "Daily" else 12
    if len(returns) < min_obs:
        return None, f"Only {len(returns)} matched {frequency.lower()} observations. Try another benchmark or shorter period."

    f_ret = returns["fund_return"].astype(float)
    b_ret = returns["benchmark_return"].astype(float)
    rf_period = float(rf_annual) / periods

    beta = float(np.cov(f_ret, b_ret, ddof=1)[0, 1] / np.var(b_ret, ddof=1)) if np.var(b_ret, ddof=1) else np.nan
    regression_intercept = float(f_ret.mean() - beta * b_ret.mean()) if not np.isnan(beta) else np.nan
    corr = float(np.corrcoef(f_ret, b_ret)[0, 1]) if len(returns) > 1 else np.nan
    r_sq = corr**2 if not np.isnan(corr) else np.nan

    fund_return_ann = float(f_ret.mean()) * periods
    benchmark_return_ann = float(b_ret.mean()) * periods
    capm_expected = float(rf_annual) + beta * (benchmark_return_ann - float(rf_annual))
    alpha = fund_return_ann - capm_expected

    fund_volatility = float(f_ret.std(ddof=1)) * np.sqrt(periods)
    benchmark_volatility = float(b_ret.std(ddof=1)) * np.sqrt(periods)
    active = f_ret - b_ret
    tracking_error = float(active.std(ddof=1)) * np.sqrt(periods)
    downside = f_ret[f_ret < rf_period] - rf_period
    downside_std = float(np.sqrt((downside**2).mean())) * np.sqrt(periods) if len(downside) else np.nan

    sharpe = (fund_return_ann - float(rf_annual)) / fund_volatility if fund_volatility else np.nan
    sortino = (fund_return_ann - float(rf_annual)) / downside_std if downside_std else np.nan
    treynor = (fund_return_ann - float(rf_annual)) / beta if beta else np.nan
    info_ratio = (fund_return_ann - benchmark_return_ann) / tracking_error if tracking_error else np.nan

    growth = prices / prices.iloc[0] * 10000
    drawdown_fund = calculate_drawdown(prices["fund"])
    drawdown_benchmark = calculate_drawdown(prices["benchmark"])
    rolling_period = periods
    rolling_fund = prices["fund"].pct_change(rolling_period)
    rolling_benchmark = prices["benchmark"].pct_change(rolling_period)

    elapsed_years = max((prices.index[-1] - prices.index[0]).days / 365.25, 1 / periods)
    live_nav, live_date, live_change, live_change_pct = fetch_live_nav(str(scheme_code))

    result = {
        "fund_name": fund_name,
        "scheme_code": str(scheme_code),
        "category": category,
        "benchmark_label": benchmark_label,
        "benchmark_source": benchmark_source,
        "frequency": frequency,
        "periods_per_year": periods,
        "period_years_requested": years,
        "elapsed_years": elapsed_years,
        "rf_annual": float(rf_annual),
        "rf_period": rf_period,
        "prices": prices,
        "returns": returns,
        "fund_return_ann": fund_return_ann,
        "benchmark_return_ann": benchmark_return_ann,
        "fund_cagr": cagr(prices["fund"]),
        "benchmark_cagr": cagr(prices["benchmark"]),
        "capm_expected": capm_expected,
        "alpha": alpha,
        "beta": beta,
        "regression_intercept": regression_intercept,
        "r_sq": r_sq,
        "correlation": corr,
        "fund_volatility": fund_volatility,
        "benchmark_volatility": benchmark_volatility,
        "tracking_error": tracking_error,
        "downside_deviation": downside_std,
        "sharpe": sharpe,
        "sortino": sortino,
        "treynor": treynor,
        "information_ratio": info_ratio,
        "active_return": fund_return_ann - benchmark_return_ann,
        "max_drawdown": float(drawdown_fund.min()) if not drawdown_fund.empty else np.nan,
        "benchmark_max_drawdown": float(drawdown_benchmark.min()) if not drawdown_benchmark.empty else np.nan,
        "growth": growth,
        "drawdown_fund": drawdown_fund,
        "drawdown_benchmark": drawdown_benchmark,
        "rolling_fund": rolling_fund,
        "rolling_benchmark": rolling_benchmark,
        "n_obs": len(returns),
        "start_date": prices.index[0],
        "end_date": prices.index[-1],
        "live_nav": live_nav,
        "live_date": live_date,
        "live_change": live_change,
        "live_change_pct": live_change_pct,
    }
    return result, None


def search_fund_widget(key_prefix, label="Fund", direct_growth_default=True):
    try:
        schemes = load_schemes()
    except Exception as exc:
        st.error(f"AMFI scheme list could not be loaded: {exc}")
        return None

    c1, c2 = st.columns([3, 1])
    with c1:
        query = st.text_input(label, placeholder="Type fund name or AMFI code", key=f"{key_prefix}_query")
    with c2:
        direct_growth = st.checkbox("Direct Growth only", value=direct_growth_default, key=f"{key_prefix}_direct")

    filtered = schemes.copy()
    if direct_growth:
        filtered = filtered[
            filtered["name_lower"].str.contains("direct", na=False)
            & filtered["name_lower"].str.contains("growth", na=False)
            & ~filtered["name_lower"].str.contains("idcw|dividend|payout|bonus", na=False)
        ]
    tokens = [t.strip().lower() for t in re.split(r"\s+", query or "") if t.strip()]
    for token in tokens:
        filtered = filtered[filtered["name_lower"].str.contains(re.escape(token), na=False) | filtered["code"].str.contains(token, na=False)]
    filtered = filtered.head(75)
    if filtered.empty:
        st.info("No matching fund found. Try fewer words or turn off Direct Growth only.")
        return None

    options = [f"{row.name} | {row.code}" for row in filtered.itertuples(index=False)]
    pick = st.selectbox("Select match", options, key=f"{key_prefix}_select")
    name, code = pick.rsplit("|", 1)
    name = name.strip()
    code = code.strip()
    return {"code": code, "name": name, "category": category_from_name(name)}


def render_header():
    st.markdown(
        f"""
<div class="topbar">
  <div>
    <div class="brand-title">{APP_TITLE}</div>
    <div class="brand-sub">Excel-style CAPM calculations, clean fund comparison, scenario forecasting, and premium workbook export. No AI, no unreliable allocation estimates.</div>
  </div>
  <div class="status-pill">Formula mode: Excel aligned</div>
</div>
""",
        unsafe_allow_html=True,
    )
    page = st.radio("Page", PAGES, horizontal=True, label_visibility="collapsed")
    return page


def render_analysis_inputs():
    section("Analysis Setup")
    fund = search_fund_widget("main", "Search primary fund")
    a, b, c, d = st.columns(4)
    with a:
        benchmark_choice = st.selectbox("Benchmark", list(BENCHMARKS.keys()), index=0)
    with b:
        years = st.select_slider("Period", options=[1, 2, 3, 4, 5, 7, 10], value=4)
    with c:
        frequency = st.radio("Return frequency", ["Daily", "Monthly"], horizontal=True)
    with d:
        rf_annual = st.number_input("Risk-free rate (%)", min_value=0.0, max_value=15.0, value=6.5, step=0.1) / 100
    uploaded = st.file_uploader("Optional benchmark CSV/XLSX with Date and Close/Price columns", type=["csv", "xlsx", "xls"])
    run = st.button("Run Analysis", use_container_width=True)
    return fund, benchmark_choice, years, frequency, rf_annual, uploaded, run


def chart_growth(r):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=r["growth"].index, y=r["growth"]["fund"], name=r["fund_name"][:32], line=dict(color=TEAL, width=2.6)))
    fig.add_trace(go.Scatter(x=r["growth"].index, y=r["growth"]["benchmark"], name=r["benchmark_label"], line=dict(color=BLUE, width=2.1, dash="dash")))
    fig.update_layout(
        title="Growth of INR 10,000",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=355,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(gridcolor=GRID, color=MUTED),
        yaxis=dict(gridcolor=GRID, color=MUTED, tickprefix="INR "),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def chart_drawdown(r):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=r["drawdown_fund"].index, y=r["drawdown_fund"] * 100, name="Fund", line=dict(color=RED, width=2.1), fill="tozeroy", fillcolor="rgba(180,35,24,0.08)"))
    fig.add_trace(go.Scatter(x=r["drawdown_benchmark"].index, y=r["drawdown_benchmark"] * 100, name=r["benchmark_label"], line=dict(color=GOLD, width=1.8, dash="dash")))
    fig.update_layout(
        title="Drawdown",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=325,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(gridcolor=GRID, color=MUTED),
        yaxis=dict(gridcolor=GRID, color=MUTED, ticksuffix="%"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def chart_rolling(r):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=r["rolling_fund"].dropna().index, y=r["rolling_fund"].dropna() * 100, name="Fund", line=dict(color=TEAL, width=2.2)))
    fig.add_trace(go.Scatter(x=r["rolling_benchmark"].dropna().index, y=r["rolling_benchmark"].dropna() * 100, name=r["benchmark_label"], line=dict(color=BLUE, width=1.9, dash="dash")))
    fig.add_hline(y=0, line_dash="dot", line_color="#98A2B3")
    fig.update_layout(
        title=f"Rolling 1Y Return ({r['frequency']})",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=325,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(gridcolor=GRID, color=MUTED),
        yaxis=dict(gridcolor=GRID, color=MUTED, ticksuffix="%"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def chart_capm_scatter(r):
    returns = r["returns"]
    x = returns["benchmark_return"] * 100
    y = returns["fund_return"] * 100
    x_line = np.linspace(float(x.min()), float(x.max()), 100)
    y_line = (r["regression_intercept"] * 100) + r["beta"] * x_line
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y, mode="markers", name="Return observations", marker=dict(color=TEAL, size=5, opacity=0.45)))
    fig.add_trace(go.Scatter(x=x_line, y=y_line, mode="lines", name=f"Beta {num(r['beta'], 2)}", line=dict(color=BLUE, width=2.2)))
    fig.add_hline(y=0, line_dash="dot", line_color="#98A2B3")
    fig.add_vline(x=0, line_dash="dot", line_color="#98A2B3")
    fig.update_layout(
        title="CAPM Regression View",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=325,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(title="Benchmark return (%)", gridcolor=GRID, color=MUTED),
        yaxis=dict(title="Fund return (%)", gridcolor=GRID, color=MUTED),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
    )
    st.plotly_chart(fig, use_container_width=True)


def render_result_summary(r):
    st.markdown(
        f"""
<div class="page-note">
  <b>{esc(r['fund_name'])}</b> | Code {esc(r['scheme_code'])} | {category_label(r['category'])} |
  Benchmark: {esc(r['benchmark_label'])} ({esc(r['benchmark_source'])}) |
  {r['n_obs']} matched {r['frequency'].lower()} observations from {r['start_date'].strftime('%d %b %Y')} to {r['end_date'].strftime('%d %b %Y')}.
</div>
""",
        unsafe_allow_html=True,
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(metric_card("Excel-style annual return", pct(r["fund_return_ann"]), f"Benchmark {pct(r['benchmark_return_ann'])}", signed_class(r["fund_return_ann"])), unsafe_allow_html=True)
    c2.markdown(metric_card("Jensen alpha", pct(r["alpha"]), "Fund return minus CAPM expected", signed_class(r["alpha"])), unsafe_allow_html=True)
    c3.markdown(metric_card("Beta", num(r["beta"], 2), "Sensitivity to benchmark", "bad" if r["beta"] > 1.1 else "good" if r["beta"] < 0.9 else "warn"), unsafe_allow_html=True)
    c4.markdown(metric_card("Sharpe ratio", num(r["sharpe"], 2), "Excess return per unit risk", "good" if r["sharpe"] > 1 else "warn"), unsafe_allow_html=True)


def render_result_tabs(r):
    overview, performance, risk, capm, charts, data = st.tabs(["Overview", "Performance", "Risk", "CAPM", "Charts", "Data"])
    with overview:
        a, b, c = st.columns(3)
        with a:
            st.markdown(
                "<div class='panel'><div class='card-title'>Current NAV</div>"
                + row_html("Live NAV", f"INR {num(r['live_nav'], 4)}" if r["live_nav"] else "N/A")
                + row_html("NAV date", r["live_date"] or "N/A")
                + row_html("Last change", f"{num(r['live_change'], 4)} ({pct(r['live_change_pct'])})", signed_class(r["live_change_pct"]))
                + "</div>",
                unsafe_allow_html=True,
            )
        with b:
            st.markdown(
                "<div class='panel'><div class='card-title'>Return Snapshot</div>"
                + row_html("Annual return", pct(r["fund_return_ann"]), signed_class(r["fund_return_ann"]))
                + row_html("CAGR", pct(r["fund_cagr"]), signed_class(r["fund_cagr"]))
                + row_html("Active return", pct(r["active_return"]), signed_class(r["active_return"]))
                + "</div>",
                unsafe_allow_html=True,
            )
        with c:
            st.markdown(
                "<div class='panel'><div class='card-title'>Risk Snapshot</div>"
                + row_html("Volatility", pct(r["fund_volatility"]))
                + row_html("Max drawdown", pct(r["max_drawdown"]), "bad")
                + row_html("Tracking error", pct(r["tracking_error"]))
                + "</div>",
                unsafe_allow_html=True,
            )
    with performance:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                "<div class='panel'><div class='card-title'>Fund vs Benchmark</div>"
                + row_html("Fund annual return", pct(r["fund_return_ann"]), signed_class(r["fund_return_ann"]))
                + row_html("Benchmark annual return", pct(r["benchmark_return_ann"]), signed_class(r["benchmark_return_ann"]))
                + row_html("Fund CAGR", pct(r["fund_cagr"]), signed_class(r["fund_cagr"]))
                + row_html("Benchmark CAGR", pct(r["benchmark_cagr"]), signed_class(r["benchmark_cagr"]))
                + row_html("Active annual return", pct(r["active_return"]), signed_class(r["active_return"]))
                + "</div>",
                unsafe_allow_html=True,
            )
        with c2:
            chart_growth(r)
    with risk:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                "<div class='panel'><div class='card-title'>Risk Metrics</div>"
                + row_html("Fund volatility", pct(r["fund_volatility"]))
                + row_html("Benchmark volatility", pct(r["benchmark_volatility"]))
                + row_html("Downside deviation", pct(r["downside_deviation"]))
                + row_html("Max drawdown", pct(r["max_drawdown"]), "bad")
                + row_html("Benchmark max drawdown", pct(r["benchmark_max_drawdown"]), "bad")
                + "</div>",
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                "<div class='panel'><div class='card-title'>Risk-adjusted Ratios</div>"
                + row_html("Sharpe", num(r["sharpe"], 3), "good" if r["sharpe"] > 1 else "warn")
                + row_html("Sortino", num(r["sortino"], 3), "good" if r["sortino"] > 1 else "warn")
                + row_html("Treynor", num(r["treynor"], 3), signed_class(r["treynor"]))
                + row_html("Information ratio", num(r["information_ratio"], 3), signed_class(r["information_ratio"]))
                + row_html("Tracking error", pct(r["tracking_error"]))
                + "</div>",
                unsafe_allow_html=True,
            )
    with capm:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                "<div class='panel'><div class='card-title'>CAPM Formula Output</div>"
                + row_html("Beta", num(r["beta"], 4))
                + row_html("Risk-free rate", pct(r["rf_annual"]))
                + row_html("Market return", pct(r["benchmark_return_ann"]))
                + row_html("CAPM expected return", pct(r["capm_expected"]))
                + row_html("Jensen alpha", pct(r["alpha"]), signed_class(r["alpha"]))
                + row_html("R-squared", num(r["r_sq"], 3))
                + "</div>",
                unsafe_allow_html=True,
            )
        with c2:
            chart_capm_scatter(r)
    with charts:
        c1, c2 = st.columns(2)
        with c1:
            chart_drawdown(r)
        with c2:
            chart_rolling(r)
    with data:
        table = r["returns"].copy()
        table.insert(0, "Date", table.index)
        table["active_return"] = table["fund_return"] - table["benchmark_return"]
        st.dataframe(table, use_container_width=True, hide_index=True)


def peer_candidates(category, selected_code, selected_name, limit=8):
    try:
        schemes = load_schemes()
    except Exception:
        return []
    keywords = CATEGORY_KEYWORDS.get(category, [])
    mask = pd.Series(False, index=schemes.index)
    for keyword in keywords:
        mask = mask | schemes["name_lower"].str.contains(keyword, na=False)
    mask = mask & schemes["name_lower"].str.contains("direct", na=False)
    mask = mask & schemes["name_lower"].str.contains("growth", na=False)
    mask = mask & ~schemes["name_lower"].str.contains("idcw|dividend|regular|bonus", na=False)
    mask = mask & (schemes["code"] != str(selected_code))
    selected_amc = selected_name.split()[0].lower() if selected_name else ""
    candidates = schemes[mask].copy()
    if selected_amc:
        candidates = candidates[~candidates["name_lower"].str.startswith(selected_amc)]
    candidates["name_len"] = candidates["name"].str.len()
    candidates = candidates.sort_values(["name_len", "name"]).head(limit)
    return list(candidates[["code", "name"]].itertuples(index=False, name=None))


def render_peers(r):
    section("Comparable Funds")
    if r["benchmark_source"] == "custom":
        st.info("Peer ranking is skipped for uploaded custom benchmarks so the app does not compare peers on a different hidden benchmark.")
        return
    peers = peer_candidates(r["category"], r["scheme_code"], r["fund_name"], limit=6)
    if not peers:
        st.info("No comparable Direct Growth peers were found from the AMFI list.")
        return
    rows = []
    with st.spinner("Calculating peer metrics on the same benchmark basis..."):
        for code, name in peers:
            result, err = run_capm_analysis(code, name, category_from_name(name), r["benchmark_label"], min(r["period_years_requested"], 5), r["frequency"], r["rf_annual"])
            if result and not err:
                rows.append(
                    {
                        "Fund": name[:60],
                        "Code": code,
                        "Return": pct(result["fund_return_ann"]),
                        "CAGR": pct(result["fund_cagr"]),
                        "Alpha": pct(result["alpha"]),
                        "Beta": num(result["beta"], 2),
                        "Sharpe": num(result["sharpe"], 2),
                        "Max DD": pct(result["max_drawdown"]),
                    }
                )
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("Peer metrics could not be calculated right now because data was unavailable.")


def build_report_workbook(r):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    colors = {
        "navy": "173B57",
        "teal": "0F766E",
        "soft": "EEF4F7",
        "border": "D9E2EC",
        "text": "182230",
        "muted": "667085",
        "green": "15803D",
        "red": "B42318",
        "white": "FFFFFF",
    }
    thin = Side(style="thin", color=colors["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def style_sheet(sheet):
        sheet.sheet_view.showGridLines = False
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 18

    def write_title(sheet, title, subtitle=""):
        sheet.merge_cells("A1:H1")
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, color=colors["white"], size=16)
        sheet["A1"].fill = fill(colors["navy"])
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A2:H2")
        sheet["A2"] = subtitle
        sheet["A2"].font = Font(color=colors["muted"], italic=True)
        sheet["A2"].fill = fill(colors["soft"])

    style_sheet(ws)
    write_title(ws, "CAPM Mutual Fund Research Report", f"{r['fund_name']} | Code {r['scheme_code']} | Generated {datetime.now().strftime('%d %b %Y')}")

    summary_rows = [
        ("Fund", r["fund_name"], "Selected AMFI scheme"),
        ("Category", category_label(r["category"]), "Detected from fund name"),
        ("Benchmark", r["benchmark_label"], r["benchmark_source"]),
        ("Period", f"{r['start_date'].strftime('%d %b %Y')} to {r['end_date'].strftime('%d %b %Y')}", f"{r['n_obs']} {r['frequency'].lower()} observations"),
        ("Excel-style annual return", pct(r["fund_return_ann"]), "Average periodic return annualized"),
        ("Benchmark annual return", pct(r["benchmark_return_ann"]), "Same frequency and matched dates"),
        ("CAPM expected return", pct(r["capm_expected"]), "Rf + Beta x market premium"),
        ("Jensen alpha", pct(r["alpha"]), "Annual return minus CAPM expected"),
        ("Beta", num(r["beta"], 4), "SLOPE equivalent"),
        ("Sharpe", num(r["sharpe"], 4), "Risk-adjusted return"),
        ("Sortino", num(r["sortino"], 4), "Downside-risk adjusted return"),
        ("Max drawdown", pct(r["max_drawdown"]), "Worst fall from running peak"),
    ]
    for row_idx, row in enumerate(summary_rows, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.fill = fill(colors["white"] if row_idx % 2 else colors["soft"])
            cell.font = Font(color=colors["text"], bold=(col_idx == 1))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 48

    dash = wb.create_sheet("Metrics Dashboard")
    style_sheet(dash)
    write_title(dash, "Metrics Dashboard", "Primary metrics for investment research review")
    dashboard_rows = [
        ("Return", "Fund Annual Return", r["fund_return_ann"]),
        ("Return", "Fund CAGR", r["fund_cagr"]),
        ("Return", "Active Return", r["active_return"]),
        ("CAPM", "CAPM Expected", r["capm_expected"]),
        ("CAPM", "Jensen Alpha", r["alpha"]),
        ("CAPM", "Beta", r["beta"]),
        ("Risk", "Volatility", r["fund_volatility"]),
        ("Risk", "Tracking Error", r["tracking_error"]),
        ("Risk", "Max Drawdown", r["max_drawdown"]),
        ("Ratio", "Sharpe", r["sharpe"]),
        ("Ratio", "Sortino", r["sortino"]),
        ("Ratio", "Treynor", r["treynor"]),
        ("Ratio", "Information Ratio", r["information_ratio"]),
    ]
    dash.append([])
    headers = ["Bucket", "Metric", "Value", "Display"]
    for col, h in enumerate(headers, 1):
        cell = dash.cell(4, col, h)
        cell.font = Font(bold=True, color=colors["white"])
        cell.fill = fill(colors["teal"])
        cell.border = border
    for idx, (bucket, metric, value) in enumerate(dashboard_rows, 5):
        dash.cell(idx, 1, bucket)
        dash.cell(idx, 2, metric)
        dash.cell(idx, 3, float(value) if pd.notna(value) else "")
        dash.cell(idx, 4, pct(value) if metric not in ["Beta", "Sharpe", "Sortino", "Treynor", "Information Ratio"] else num(value, 4))
        for col in range(1, 5):
            cell = dash.cell(idx, col)
            cell.border = border
            cell.fill = fill(colors["white"] if idx % 2 else colors["soft"])
            if col == 3:
                cell.number_format = "0.0000"

    assumptions = wb.create_sheet("Assumptions")
    style_sheet(assumptions)
    write_title(assumptions, "Assumptions", "Data and calculation settings used in this report")
    assumption_rows = [
        ("Data source", "AMFI NAV via mftool, benchmark via Yahoo Finance or uploaded file"),
        ("Date alignment", "Only matched dates with both fund and benchmark observations are used"),
        ("Annualization", f"{r['periods_per_year']} periods per year for {r['frequency']} returns"),
        ("Risk-free rate", pct(r["rf_annual"])),
        ("Asset allocation", "Not included because AMFI NAV history does not provide reliable holdings/allocation"),
        ("AI commentary", "Not included by design"),
    ]
    for idx, row in enumerate(assumption_rows, 4):
        assumptions.cell(idx, 1, row[0])
        assumptions.cell(idx, 2, row[1])
        for col in range(1, 3):
            cell = assumptions.cell(idx, col)
            cell.border = border
            cell.fill = fill(colors["white"] if idx % 2 else colors["soft"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    assumptions.column_dimensions["A"].width = 28
    assumptions.column_dimensions["B"].width = 90

    audit = wb.create_sheet("Formula Audit")
    style_sheet(audit)
    write_title(audit, "Formula Audit", "Excel-style formula reference")
    formula_rows = formula_audit_rows(r["frequency"])
    for col, h in enumerate(["Metric", "Formula", "Notes"], 1):
        cell = audit.cell(4, col, h)
        cell.font = Font(bold=True, color=colors["white"])
        cell.fill = fill(colors["teal"])
        cell.border = border
    for idx, row in enumerate(formula_rows, 5):
        audit.cell(idx, 1, row["Metric"])
        audit.cell(idx, 2, row["Formula"])
        audit.cell(idx, 3, row["Notes"])
        for col in range(1, 4):
            cell = audit.cell(idx, col)
            cell.border = border
            cell.fill = fill(colors["white"] if idx % 2 else colors["soft"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    audit.column_dimensions["A"].width = 24
    audit.column_dimensions["B"].width = 58
    audit.column_dimensions["C"].width = 58

    raw = wb.create_sheet("Return Data")
    raw.sheet_view.showGridLines = False
    raw_headers = ["Date", "Fund Price", "Benchmark Price", "Fund Return", "Benchmark Return", "Active Return", "Fund Growth", "Benchmark Growth", "Fund Drawdown", "Benchmark Drawdown"]
    for col, h in enumerate(raw_headers, 1):
        cell = raw.cell(1, col, h)
        cell.font = Font(bold=True, color=colors["white"])
        cell.fill = fill(colors["navy"])
        cell.border = border
        raw.column_dimensions[get_column_letter(col)].width = 18

    active = r["returns"]["fund_return"] - r["returns"]["benchmark_return"]
    data = pd.DataFrame(
        {
            "Date": r["prices"].index,
            "Fund Price": r["prices"]["fund"].values,
            "Benchmark Price": r["prices"]["benchmark"].values,
            "Fund Return": r["returns"]["fund_return"].reindex(r["prices"].index).values,
            "Benchmark Return": r["returns"]["benchmark_return"].reindex(r["prices"].index).values,
            "Active Return": active.reindex(r["prices"].index).values,
            "Fund Growth": r["growth"]["fund"].values,
            "Benchmark Growth": r["growth"]["benchmark"].values,
            "Fund Drawdown": r["drawdown_fund"].reindex(r["prices"].index).values,
            "Benchmark Drawdown": r["drawdown_benchmark"].reindex(r["prices"].index).values,
        }
    )
    for ridx, row in enumerate(data.itertuples(index=False), 2):
        for cidx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            cell = raw.cell(ridx, cidx, value)
            cell.border = border
            if cidx == 1 and hasattr(value, "strftime"):
                cell.value = value.strftime("%Y-%m-%d")
            if cidx in [4, 5, 6, 9, 10]:
                cell.number_format = "0.00%"
            elif cidx > 1:
                cell.number_format = "0.00"

    if len(data) > 2:
        chart = LineChart()
        chart.title = "Growth of INR 10,000"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Date"
        chart.add_data(Reference(raw, min_col=7, max_col=8, min_row=1, max_row=min(len(data) + 1, 260)), titles_from_data=True)
        chart.set_categories(Reference(raw, min_col=1, min_row=2, max_row=min(len(data) + 1, 260)))
        chart.height = 8
        chart.width = 18
        dash.add_chart(chart, "F4")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def formula_audit_rows(frequency="Daily"):
    periods = 252 if frequency == "Daily" else 12
    return [
        {"Metric": "Periodic Return", "Formula": "Price[t] / Price[t-1] - 1", "Notes": "Used for both fund NAV and benchmark price after date alignment."},
        {"Metric": "Annual Return", "Formula": f"AVERAGE(periodic returns) x {periods}", "Notes": "Matches your Excel CAPM sheet style. CAGR is also shown separately."},
        {"Metric": "CAGR", "Formula": "(Ending price / Starting price)^(1 / years) - 1", "Notes": "Useful investor return view, but not the primary CAPM input."},
        {"Metric": "Beta", "Formula": "COVARIANCE.S(fund returns, benchmark returns) / VAR.S(benchmark returns)", "Notes": "Equivalent to Excel SLOPE(fund returns, benchmark returns)."},
        {"Metric": "CAPM Expected", "Formula": "Risk-free rate + Beta x (Market annual return - Risk-free rate)", "Notes": "Uses annualized average market return."},
        {"Metric": "Jensen Alpha", "Formula": "Fund annual return - CAPM expected return", "Notes": "Positive alpha means fund beat CAPM expectation over the selected window."},
        {"Metric": "Volatility", "Formula": f"STDEV.S(periodic returns) x SQRT({periods})", "Notes": "Annualized total risk."},
        {"Metric": "Sharpe", "Formula": "(Fund annual return - Risk-free rate) / Volatility", "Notes": "Return per unit of total volatility."},
        {"Metric": "Sortino", "Formula": "(Fund annual return - Risk-free rate) / Downside deviation", "Notes": "Downside deviation uses returns below periodic risk-free return."},
        {"Metric": "Treynor", "Formula": "(Fund annual return - Risk-free rate) / Beta", "Notes": "Return per unit of market risk."},
        {"Metric": "Tracking Error", "Formula": f"STDEV.S(fund return - benchmark return) x SQRT({periods})", "Notes": "Measures active risk."},
        {"Metric": "Information Ratio", "Formula": "(Fund annual return - Market annual return) / Tracking error", "Notes": "Active return per unit of active risk."},
        {"Metric": "Max Drawdown", "Formula": "(Current value - Running peak) / Running peak", "Notes": "Worst observed drawdown over the selected period."},
    ]


def render_download_report(r, key="report"):
    report = build_report_workbook(r)
    st.download_button(
        "Download Premium Excel Report",
        data=report,
        file_name=f"CAPM_Research_{r['scheme_code']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_{key}",
        use_container_width=True,
    )


def page_analyze():
    st.markdown("<div class='page-note'>Start here. Search a fund, choose a reliable benchmark, and run an Excel-style CAPM analysis. Asset allocation and AI have been removed deliberately.</div>", unsafe_allow_html=True)
    fund, benchmark_choice, years, frequency, rf_annual, uploaded, run = render_analysis_inputs()
    if run:
        if not fund:
            st.warning("Please select a fund before running analysis.")
            return
        with st.spinner("Running Excel-style CAPM calculation..."):
            result, err = run_capm_analysis(fund["code"], fund["name"], fund["category"], benchmark_choice, years, frequency, rf_annual, uploaded)
        if err:
            st.error(err)
            return
        st.session_state["last_analysis"] = result
    r = st.session_state.get("last_analysis")
    if not r:
        return
    section("Analysis Result")
    render_result_summary(r)
    render_result_tabs(r)
    render_peers(r)
    section("Export")
    render_download_report(r, "analyze")


def page_compare():
    st.markdown("<div class='page-note'>Compare funds on one benchmark and one calculation basis. This keeps the comparison fair and avoids mixed assumptions.</div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        fund_a = search_fund_widget("compare_a", "First fund")
    with c2:
        fund_b = search_fund_widget("compare_b", "Second fund")
    a, b, c = st.columns(3)
    with a:
        benchmark_choice = st.selectbox("Benchmark", list(BENCHMARKS.keys()), index=0, key="compare_bench")
    with b:
        years = st.select_slider("Period", options=[1, 2, 3, 4, 5, 7, 10], value=4, key="compare_years")
    with c:
        frequency = st.radio("Frequency", ["Daily", "Monthly"], horizontal=True, key="compare_freq")
    rf_annual = st.number_input("Risk-free rate (%)", min_value=0.0, max_value=15.0, value=6.5, step=0.1, key="compare_rf") / 100
    run = st.button("Compare Funds", use_container_width=True)

    if not run:
        return
    if not fund_a or not fund_b:
        st.warning("Select both funds before comparing.")
        return

    results = []
    with st.spinner("Calculating both funds on matched assumptions..."):
        for fund in [fund_a, fund_b]:
            result, err = run_capm_analysis(fund["code"], fund["name"], fund["category"], benchmark_choice, years, frequency, rf_annual)
            if err:
                st.warning(f"{fund['name'][:50]}: {err}")
            else:
                results.append(result)
    if len(results) < 2:
        return

    section("Comparison Table")
    rows = []
    for r in results:
        rows.append(
            {
                "Fund": r["fund_name"][:60],
                "Category": category_label(r["category"]),
                "Annual Return": pct(r["fund_return_ann"]),
                "CAGR": pct(r["fund_cagr"]),
                "Alpha": pct(r["alpha"]),
                "Beta": num(r["beta"], 2),
                "Sharpe": num(r["sharpe"], 2),
                "Sortino": num(r["sortino"], 2),
                "Max DD": pct(r["max_drawdown"]),
                "Tracking Error": pct(r["tracking_error"]),
            }
        )
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    section("Growth Comparison")
    fig = go.Figure()
    for idx, r in enumerate(results):
        fig.add_trace(go.Scatter(x=r["growth"].index, y=r["growth"]["fund"], name=r["fund_name"][:32], line=dict(color=[TEAL, ORANGE][idx], width=2.4)))
    fig.update_layout(
        title="Growth of INR 10,000",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=360,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(gridcolor=GRID, color=MUTED),
        yaxis=dict(gridcolor=GRID, color=MUTED, tickprefix="INR "),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def scenario_projection(current, monthly_sip, annual_rate, years):
    months = int(years * 12)
    rm = (1 + annual_rate) ** (1 / 12) - 1
    future_current = current * ((1 + rm) ** months)
    if abs(rm) < 1e-9:
        future_sip = monthly_sip * months
    else:
        future_sip = monthly_sip * (((1 + rm) ** months - 1) / rm) * (1 + rm)
    return future_current + future_sip


def page_forecast():
    st.markdown("<div class='page-note'>Forecasting is scenario planning, not prediction. It uses the latest analysis return and volatility when available, or your manual expected return.</div>", unsafe_allow_html=True)
    last = st.session_state.get("last_analysis")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        current = st.number_input("Existing investment", min_value=0, value=0, step=10000)
    with c2:
        sip = st.number_input("Monthly SIP", min_value=0, value=10000, step=1000)
    with c3:
        goal = st.number_input("Goal amount", min_value=0, value=1000000, step=50000)
    with c4:
        years = st.slider("Horizon years", 1, 30, 7)

    if last:
        base_return = float(last["fund_cagr"] if pd.notna(last["fund_cagr"]) else last["fund_return_ann"])
        vol = float(last["fund_volatility"] if pd.notna(last["fund_volatility"]) else 0.12)
        st.info(f"Using latest analyzed fund: {last['fund_name']} | Base return {pct(base_return)} | Volatility {pct(vol)}")
    else:
        base_return = st.number_input("Expected annual return (%)", min_value=-10.0, max_value=30.0, value=12.0, step=0.5) / 100
        vol = st.number_input("Expected volatility (%)", min_value=0.0, max_value=50.0, value=15.0, step=0.5) / 100

    vol_penalty = min(vol * 0.45, 0.12)
    scenarios = {
        "Conservative": max(base_return - vol_penalty, -0.08),
        "Base": base_return,
        "Optimistic": min(base_return + vol_penalty, 0.30),
    }
    rows = []
    for name, rate in scenarios.items():
        value = scenario_projection(current, sip, rate, years)
        rows.append(
            {
                "Scenario": name,
                "Annual Return Used": pct(rate),
                "Projected Value": money(value),
                "Goal Gap": money(value - goal),
                "Goal Met": "Yes" if value >= goal else "No",
            }
        )
    section("Forecast Result")
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    fig = go.Figure()
    for name, rate in scenarios.items():
        xs = list(range(0, years + 1))
        ys = [scenario_projection(current, sip, rate, y) for y in xs]
        fig.add_trace(go.Scatter(x=xs, y=ys, name=name, mode="lines+markers"))
    fig.add_hline(y=goal, line_dash="dash", line_color=RED, annotation_text="Goal")
    fig.update_layout(
        title="Scenario Projection",
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        height=380,
        margin=dict(l=10, r=10, t=45, b=10),
        xaxis=dict(title="Years", gridcolor=GRID, color=MUTED),
        yaxis=dict(title="Projected value", gridcolor=GRID, color=MUTED, tickprefix="INR "),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def page_report_builder():
    st.markdown("<div class='page-note'>The report builder exports a research-style workbook, not a raw dump. Run an analysis first, then export from here.</div>", unsafe_allow_html=True)
    r = st.session_state.get("last_analysis")
    if not r:
        st.info("No analysis is available yet. Go to Analyze Fund and run one fund first.")
        return
    render_result_summary(r)
    section("Workbook Contents")
    st.markdown(
        """
<div class="panel">
  <div class="card-title">Included sheets</div>
  <div class="row"><span>Executive Summary</span><span>Fund, benchmark, return, risk, CAPM summary</span></div>
  <div class="row"><span>Metrics Dashboard</span><span>Clean metric table plus growth chart</span></div>
  <div class="row"><span>Assumptions</span><span>Data source, frequency, risk-free rate, exclusions</span></div>
  <div class="row"><span>Formula Audit</span><span>Formula-by-formula calculation explanation</span></div>
  <div class="row"><span>Return Data</span><span>Aligned raw prices, returns, growth, drawdown</span></div>
</div>
""",
        unsafe_allow_html=True,
    )
    render_download_report(r, "builder")


def page_formula_audit():
    st.markdown("<div class='page-note'>This page documents the calculation backend. It is intentionally transparent so the app can be checked against your Excel workbook.</div>", unsafe_allow_html=True)
    frequency = st.radio("Audit frequency", ["Daily", "Monthly"], horizontal=True)
    rows = formula_audit_rows(frequency)
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    section("Removed Features")
    st.markdown(
        """
<div class="panel">
  <div class="row"><span>AI analysis</span><span>Removed. It can be added later with secrets/env input.</span></div>
  <div class="row"><span>Asset allocation</span><span>Removed. AMFI NAV history does not provide reliable allocation or holdings.</span></div>
  <div class="row"><span>News/factsheet scraping</span><span>Removed from v1 rebuild to keep the app reliable.</span></div>
  <div class="row"><span>Sidebar</span><span>Removed. All workflows are page-based.</span></div>
</div>
""",
        unsafe_allow_html=True,
    )


def main():
    page = render_header()
    if page == "Analyze Fund":
        page_analyze()
    elif page == "Compare Funds":
        page_compare()
    elif page == "Forecast Planner":
        page_forecast()
    elif page == "Report Builder":
        page_report_builder()
    elif page == "Formula Audit":
        page_formula_audit()


if __name__ == "__main__":
    main()
